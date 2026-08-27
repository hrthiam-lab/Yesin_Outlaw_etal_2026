"""
@author: ayesin

Nuclear protein intensity and nuclear shape quantification for fixed,
immunostained single cells. Protein/DNA intensities are measured from a
DAPI mask; nuclear shape is measured from an LBR mask. Cells without an
LBR channel are quantified for protein/DNA intensity only.

Requirements:
    pip install numpy scikit-image scipy tifffile openpyxl

Usage:
    python nuclear_IF_analysis.py
"""

import os
import sys
import glob
import numpy as np
from pathlib import Path

import tifffile
from skimage import filters, measure
from scipy import ndimage

try:
    import tkinter as tk
    from tkinter import filedialog, ttk
    HAS_TK = True
except ImportError:
    HAS_TK = False

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ===================== THRESHOLD METHODS ====================================
THRESHOLD_METHODS = {
    "Triangle": filters.threshold_triangle,
    "Otsu": filters.threshold_otsu,
    "Li": filters.threshold_li,
    "Yen": filters.threshold_yen,
    "Isodata": filters.threshold_isodata,
    "Mean": filters.threshold_mean,
    "Minimum": filters.threshold_minimum,
}

# Fiji's default erosion: 3x3 cross (4-connected)
FIJI_CROSS_FOOTPRINT = np.array([[0, 1, 0],
                                  [1, 1, 1],
                                  [0, 1, 0]], dtype=bool)

# Laplacian kernel for focus scoring
LAPLACIAN_KERNEL = np.array([[0, 1, 0],
                              [1, -4, 1],
                              [0, 1, 0]], dtype=np.float64)


# ===================== GUI PARAMETER DIALOG =================================
class ParameterDialog:
    def __init__(self):
        self.result = None
        self.root = tk.Tk()
        self.root.title("Image parameters")
        self.root.resizable(False, False)

        frame = ttk.Frame(self.root, padding=20)
        frame.grid()

        row = 0
        ttk.Label(frame, text="Channel for lamins (B1/B2/AC):").grid(row=row, column=0, sticky="w", pady=4)
        self.ch_lamin = ttk.Spinbox(frame, from_=1, to=8, width=8)
        self.ch_lamin.set(4)
        self.ch_lamin.grid(row=row, column=1, sticky="w", pady=4)

        row += 1
        ttk.Label(frame, text="Channel for LBR (0 = not stained in this dataset):").grid(row=row, column=0, sticky="w", pady=4)
        self.ch_lbr = ttk.Spinbox(frame, from_=0, to=8, width=8)
        self.ch_lbr.set(3)
        self.ch_lbr.grid(row=row, column=1, sticky="w", pady=4)

        row += 1
        ttk.Label(frame, text="Channel for the nucleus mask (DAPI):").grid(row=row, column=0, sticky="w", pady=4)
        self.ch_nucl = ttk.Spinbox(frame, from_=1, to=8, width=8)
        self.ch_nucl.set(2)
        self.ch_nucl.grid(row=row, column=1, sticky="w", pady=4)

        row += 1
        ttk.Label(frame, text="Thresholding method for DAPI mask:").grid(row=row, column=0, sticky="w", pady=4)
        self.thresh_method = ttk.Combobox(frame, values=list(THRESHOLD_METHODS.keys()), width=15, state="readonly")
        self.thresh_method.set("Triangle")
        self.thresh_method.grid(row=row, column=1, sticky="w", pady=4)

        row += 1
        ttk.Label(frame, text="Pixel size (um/px):").grid(row=row, column=0, sticky="w", pady=4)
        self.px_size = ttk.Entry(frame, width=12)
        self.px_size.insert(0, "0.1083")
        self.px_size.grid(row=row, column=1, sticky="w", pady=4)

        row += 1
        ttk.Label(frame, text="LBR shape percentile cutoff:").grid(row=row, column=0, sticky="w", pady=4)
        self.lbr_percentile = ttk.Entry(frame, width=12)
        self.lbr_percentile.insert(0, "10")
        self.lbr_percentile.grid(row=row, column=1, sticky="w", pady=4)

        row += 1
        ttk.Label(frame, text="LBR shape edge smoothing (sigma):").grid(row=row, column=0, sticky="w", pady=4)
        self.smooth_sigma = ttk.Entry(frame, width=12)
        self.smooth_sigma.insert(0, "1.5")
        self.smooth_sigma.grid(row=row, column=1, sticky="w", pady=4)

        row += 1
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=(15, 0))
        ttk.Button(btn_frame, text="OK", command=self._on_ok, width=10).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Cancel", command=self._on_cancel, width=10).pack(side="left", padx=5)

        self.root.protocol("WM_DELETE_WINDOW", self._on_cancel)
        self.root.mainloop()

    def _on_ok(self):
        try:
            self.result = {
                "ch_lamin": int(self.ch_lamin.get()),
                "ch_lbr": int(self.ch_lbr.get()),
                "ch_nucl": int(self.ch_nucl.get()),
                "threshold": self.thresh_method.get(),
                "pixelsize": float(self.px_size.get()),
                "lbr_percentile": float(self.lbr_percentile.get()),
                "smooth_sigma": float(self.smooth_sigma.get()),
            }
        except ValueError:
            from tkinter import messagebox
            messagebox.showerror("Invalid input", "Please enter valid numeric values.")
            return
        self.root.destroy()

    def _on_cancel(self):
        self.result = None
        self.root.destroy()


def get_params_cli():
    print("\n===== Image parameters =====")
    ch_lamin = int(input("Channel for lamins (B1/B2/AC) [4]: ") or 4)
    ch_lbr = int(input("Channel for LBR (0 = not stained in this dataset) [3]: ") or 3)
    ch_nucl = int(input("Channel for nucleus mask (DAPI) [2]: ") or 2)
    threshold = input("Thresholding method for DAPI mask [Triangle]: ") or "Triangle"
    pixelsize = float(input("Pixel size (um/px) [0.1083]: ") or 0.1083)
    lbr_percentile = float(input("LBR shape percentile cutoff [10]: ") or 10)
    smooth_sigma = float(input("LBR shape edge smoothing sigma [1.5]: ") or 1.5)
    return {
        "ch_lamin": ch_lamin, "ch_lbr": ch_lbr, "ch_nucl": ch_nucl,
        "threshold": threshold, "pixelsize": pixelsize,
        "lbr_percentile": lbr_percentile, "smooth_sigma": smooth_sigma,
    }


def select_folders():
    if HAS_TK:
        root = tk.Tk()
        root.withdraw()
        data_dir = filedialog.askdirectory(title="Select folder containing single cells to analyze")
        if not data_dir:
            sys.exit("No input folder selected.")
        save_dir = filedialog.askdirectory(title="Select saving folder")
        if not save_dir:
            sys.exit("No output folder selected.")
        root.destroy()
    else:
        data_dir = input("Path to folder containing single cells: ").strip()
        save_dir = input("Path to saving folder: ").strip()
    return data_dir, save_dir


# ===================== IMAGE PROCESSING HELPERS =============================

def z_project(stack):
    """Average-intensity projection."""
    if stack.ndim == 2:
        return stack.astype(np.float64)
    return np.mean(stack, axis=0).astype(np.float64)


def find_best_z(stack):
    """Best-focus Z-plane by Laplacian variance (higher = sharper)."""
    if stack.ndim == 2:
        return 0
    scores = []
    for z in range(stack.shape[0]):
        plane = stack[z].astype(np.float64)
        blurred = ndimage.gaussian_filter(plane, sigma=0.7)
        lap = ndimage.convolve(blurred, LAPLACIAN_KERNEL)
        scores.append(np.var(lap))
    return int(np.argmax(scores))


def n_channels(img):
    """Best-effort number of channels in an image array, using the same
    axis-guessing convention as extract_channel."""
    if img.ndim == 2:
        return 1
    if img.ndim == 3:
        return img.shape[0] if img.shape[0] <= 8 else 1
    if img.ndim == 4:
        if img.shape[1] <= 8:
            return img.shape[1]
        if img.shape[0] <= 8:
            return img.shape[0]
        return None
    if img.ndim == 5:
        if img.shape[2] <= 8:
            return img.shape[2]
        if img.shape[1] <= 8:
            return img.shape[1]
        return None
    return None


def extract_channel(img, channel_idx):
    """Extract a single channel (1-based index). Raises if it doesn't exist."""
    c = channel_idx - 1
    nc = n_channels(img)
    if nc is None:
        raise ValueError(f"Cannot determine channel axis for shape {img.shape}.")
    if c < 0 or c >= nc:
        raise ValueError(
            f"Requested channel {channel_idx} but this image only has {nc} channel(s) "
            f"(shape {img.shape})."
        )

    if img.ndim == 2:
        return img
    if img.ndim == 3:
        return img[c]
    if img.ndim == 4:
        if img.shape[1] <= 8:
            return img[:, c, :, :]
        return img[c, :, :, :]
    if img.ndim == 5:
        if img.shape[2] <= 8:
            return img[0, :, c, :, :]
        return img[0, c, :, :, :]
    raise ValueError(f"Unsupported image dimensions: {img.ndim}")


def channel_available(img, channel_idx):
    """True if channel_idx (1-based, and > 0) exists in this image."""
    if channel_idx <= 0:
        return False
    nc = n_channels(img)
    if nc is None:
        return False
    return 0 <= (channel_idx - 1) < nc


def to_8bit(img_2d):
    """Convert to 8-bit exactly like Fiji: map [min,max] to [0,255]."""
    mn, mx = img_2d.min(), img_2d.max()
    if mx == mn:
        return np.zeros(img_2d.shape, dtype=np.uint8)
    return np.round((img_2d.astype(np.float64) - mn) / (mx - mn) * 255.0).astype(np.uint8)


def threshold_image_fiji(img_8bit, method_name):
    func = THRESHOLD_METHODS.get(method_name)
    if func is None:
        raise ValueError(f"Unknown threshold method '{method_name}'.")
    return img_8bit > func(img_8bit)


def make_nuclear_mask(dapi_2d, method_name, n_erode=2):
    """DAPI mask: 8-bit -> auto threshold -> fill holes -> erode. Used for
    all protein/DNA intensity measurements."""
    img_8bit = to_8bit(dapi_2d)
    mask_bool = threshold_image_fiji(img_8bit, method_name)
    mask_bool = ndimage.binary_fill_holes(mask_bool)
    raw_mask = mask_bool.copy()
    eroded = mask_bool.copy()
    for _ in range(n_erode):
        eroded = ndimage.binary_erosion(eroded, structure=FIJI_CROSS_FOOTPRINT)
    return raw_mask, eroded


def subtract_background(proj_2d, nuc_mask_bool):
    """Subtract mean background (pixels outside the mask). Returns (image, bg)."""
    bg_mask = ~nuc_mask_bool
    if not bg_mask.any():
        return proj_2d.copy(), 0.0
    bg_mean = float(np.mean(proj_2d[bg_mask]))
    subtracted = np.clip(proj_2d - bg_mean, 0, None)
    return subtracted, bg_mean


def lbr_outer_shape(lbr_2d, dapi_mask_bool, lbr_percentile=10, smooth_sigma=1.5):
    """Nuclear shape from the LBR best-focus plane: percentile threshold
    within the DAPI mask, fill holes, keep largest region, smooth,
    re-threshold, keep largest region again. Returns (mask, regionprops)."""
    if not dapi_mask_bool.any():
        return np.zeros_like(dapi_mask_bool), None

    lbr_mask_vals = lbr_2d[dapi_mask_bool]
    thresh_val = np.percentile(lbr_mask_vals, lbr_percentile)

    lbr_mask = (lbr_2d > thresh_val) & dapi_mask_bool
    lbr_filled = ndimage.binary_fill_holes(lbr_mask)
    lbr_filled = lbr_filled & dapi_mask_bool

    labeled = measure.label(lbr_filled.astype(np.uint8))
    props = measure.regionprops(labeled)
    if not props:
        return lbr_filled, None
    largest = max(props, key=lambda p: p.area)
    lbr_shape = (labeled == largest.label)

    if smooth_sigma > 0:
        smoothed = ndimage.gaussian_filter(lbr_shape.astype(np.float64), sigma=smooth_sigma)
        lbr_shape = (smoothed > 0.5) & dapi_mask_bool

        labeled = measure.label(lbr_shape.astype(np.uint8))
        props = measure.regionprops(labeled)
        if not props:
            return lbr_shape, None
        largest = max(props, key=lambda p: p.area)
        lbr_shape = (labeled == largest.label)

    return lbr_shape, largest


def get_prop(props, attr):
    """Get regionprops attribute, handling old/new scikit-image naming."""
    try:
        return getattr(props, attr)
    except AttributeError:
        old_map = {
            "axis_major_length": "major_axis_length",
            "axis_minor_length": "minor_axis_length",
        }
        return getattr(props, old_map.get(attr, attr))


# ===================== MAIN PIPELINE ========================================

def process_single_image(filepath, params, save_dir):
    img = tifffile.imread(filepath)
    basename = Path(filepath).stem

    ch_lamin = params["ch_lamin"]
    ch_lbr = params["ch_lbr"]
    ch_nucl = params["ch_nucl"]
    method = params["threshold"]
    px = params["pixelsize"]
    lbr_pct = params["lbr_percentile"]
    smooth_sigma = params["smooth_sigma"]

    has_lbr = channel_available(img, ch_lbr)
    if ch_lbr > 0 and not has_lbr:
        print(f"  NOTE: channel {ch_lbr} (LBR) not present in {basename} "
              f"({n_channels(img)} channel(s) found) — proceeding without LBR/shape.")

    # 1) Avg-project channels
    lamin_proj = z_project(extract_channel(img, ch_lamin))
    nucl_proj = z_project(extract_channel(img, ch_nucl))

    if has_lbr:
        lbr_stack = extract_channel(img, ch_lbr)
        lbr_proj = z_project(lbr_stack)
        best_z = find_best_z(lbr_stack)
        lbr_best_plane = lbr_stack[best_z].astype(np.float64) if lbr_stack.ndim == 3 else lbr_stack.astype(np.float64)
    else:
        lbr_proj = None
        best_z = None
        lbr_best_plane = None

    lbr_for_stack = lbr_proj if lbr_proj is not None else np.zeros_like(nucl_proj)
    proj_combined = np.stack([lamin_proj, lbr_for_stack, nucl_proj], axis=0)
    proj_path = os.path.join(save_dir, f"{basename}_Avg_Projection.tif")
    tifffile.imwrite(proj_path, proj_combined.astype(np.float32))

    # 2) DAPI nuclear mask — used for all protein/DNA intensity quantification
    raw_mask_bool, eroded_mask_bool = make_nuclear_mask(nucl_proj, method, n_erode=2)

    tifffile.imwrite(
        os.path.join(save_dir, f"{basename}_Avg_maskNucl_raw.tif"),
        (raw_mask_bool.astype(np.uint8) * 255),
    )
    tifffile.imwrite(
        os.path.join(save_dir, f"{basename}_Avg_maskNucl.tif"),
        (eroded_mask_bool.astype(np.uint8) * 255),
    )

    # 3) Binary [0,1] mask from DAPI
    bin_mask = eroded_mask_bool.astype(np.float64)
    tifffile.imwrite(
        os.path.join(save_dir, f"{basename}_Avg_maskNucl_Binary.tif"),
        bin_mask.astype(np.float32),
    )

    # 4) Background subtraction per channel
    lamin_sub, lamin_bg = subtract_background(lamin_proj, eroded_mask_bool)
    dapi_sub, dapi_bg = subtract_background(nucl_proj, eroded_mask_bool)
    if has_lbr:
        lbr_sub, lbr_bg = subtract_background(lbr_proj, eroded_mask_bool)
    else:
        lbr_sub, lbr_bg = None, np.nan

    # 5) Multiply background-subtracted channels by DAPI binary mask
    lamin_masked = lamin_sub * bin_mask
    dapi_masked = dapi_sub * bin_mask
    tifffile.imwrite(os.path.join(save_dir, f"{basename}_Avg_Avrg_Lamin_Intensity_Nuc.tif"), lamin_masked.astype(np.float32))
    tifffile.imwrite(os.path.join(save_dir, f"{basename}_Avg_Avrg_DAPI_Intensity_Nuc.tif"), dapi_masked.astype(np.float32))

    if has_lbr:
        lbr_masked = lbr_sub * bin_mask
        tifffile.imwrite(os.path.join(save_dir, f"{basename}_Avg_Avrg_LBR_Intensity_Nuc.tif"), lbr_masked.astype(np.float32))
    else:
        lbr_masked = None

    # 6) Shape from LBR best-focus Z-plane — only when present
    if has_lbr:
        lbr_shape_bool, lbr_props = lbr_outer_shape(
            lbr_best_plane, eroded_mask_bool,
            lbr_percentile=lbr_pct, smooth_sigma=smooth_sigma,
        )
        tifffile.imwrite(
            os.path.join(save_dir, f"{basename}_Avg_LBR_shape.tif"),
            (lbr_shape_bool.astype(np.uint8) * 255),
        )
        if lbr_shape_bool.any():
            outer_boundary = lbr_shape_bool.astype(np.uint8) - ndimage.binary_erosion(lbr_shape_bool).astype(np.uint8)
            tifffile.imwrite(
                os.path.join(save_dir, f"{basename}_Avg_LBR_outer_contour.tif"),
                (outer_boundary * 255).astype(np.uint8),
            )
    else:
        lbr_props = None

    # 7) Measurements

    # --- Intensity measurements (DAPI mask, background-subtracted) ---
    nuc_pix_count = int(np.sum(bin_mask))
    nuc_area_um2_dapi = nuc_pix_count * px * px

    lamin_total = float(np.sum(lamin_masked))
    dapi_total = float(np.sum(dapi_masked))
    lamin_mean = lamin_total / nuc_pix_count if nuc_pix_count > 0 else np.nan
    dapi_mean = dapi_total / nuc_pix_count if nuc_pix_count > 0 else np.nan

    if has_lbr:
        lbr_total = float(np.sum(lbr_masked))
        lbr_mean = lbr_total / nuc_pix_count if nuc_pix_count > 0 else np.nan
        lamin_over_lbr = lamin_mean / lbr_mean if (not np.isnan(lbr_mean) and lbr_mean != 0) else np.nan
    else:
        lbr_total = np.nan
        lbr_mean = np.nan
        lamin_over_lbr = np.nan

    # --- Shape measurements (LBR best-focus plane; blank if no LBR) ---
    if lbr_props is not None:
        lbr_area_px = lbr_props.area
        lbr_perim_px = lbr_props.perimeter
        lbr_major_px = get_prop(lbr_props, "axis_major_length")
        lbr_minor_px = get_prop(lbr_props, "axis_minor_length")
        lbr_solidity = lbr_props.solidity

        lbr_area_um2 = lbr_area_px * px * px
        lbr_perim_um = lbr_perim_px * px

        lbr_circ = (4.0 * np.pi * lbr_area_px) / (lbr_perim_px ** 2) if lbr_perim_px > 0 else np.nan
        lbr_round = (4.0 * lbr_area_px) / (np.pi * lbr_major_px ** 2) if lbr_major_px > 0 else np.nan
        lbr_aspect = lbr_major_px / lbr_minor_px if lbr_minor_px > 0 else np.nan
        lbr_compact = (lbr_perim_um ** 2) / (lbr_area_um2) if lbr_area_um2 > 0 else np.nan
    else:
        lbr_area_px = lbr_area_um2 = lbr_perim_um = np.nan
        lbr_circ = lbr_round = lbr_aspect = lbr_solidity = lbr_compact = np.nan

    def fmt(val, decimals=4):
        return round(val, decimals) if (val is not None and not (isinstance(val, float) and np.isnan(val))) else ""

    result = {
        "Filename": basename,
        "Has LBR": has_lbr,
        "Best Z (LBR)": (best_z + 1) if best_z is not None else "",
        # Intensity measurements (DAPI mask, BG-subtracted)
        "NucPixCount": nuc_pix_count,
        "DAPI Mask Area (um^2)": fmt(nuc_area_um2_dapi),
        "BG Lamin": round(lamin_bg, 3),
        "BG LBR": fmt(lbr_bg, 3),
        "BG DAPI": round(dapi_bg, 3),
        "Total Lamin": round(lamin_total, 3),
        "Mean Lamin": fmt(lamin_mean, 6),
        "Total LBR": fmt(lbr_total, 3),
        "Mean LBR": fmt(lbr_mean, 6),
        "Total DAPI": round(dapi_total, 3),
        "Mean DAPI": fmt(dapi_mean, 6),
        "Mean Lamin / Mean LBR": fmt(lamin_over_lbr, 6),
        # Shape measurements (LBR best-focus plane; blank if no LBR channel)
        "LBR Shape Area (px)": int(lbr_area_px) if not np.isnan(lbr_area_px) else "",
        "LBR Shape Area (um^2)": fmt(lbr_area_um2),
        "LBR Shape Perimeter (um)": fmt(lbr_perim_um),
        "LBR Shape Circularity": fmt(lbr_circ),
        "LBR Shape Roundness": fmt(lbr_round),
        "LBR Shape Solidity": fmt(lbr_solidity),
        "LBR Shape Aspect Ratio": fmt(lbr_aspect),
        "LBR Shape Compactness": fmt(lbr_compact),
    }

    bg_str = f"{lamin_bg:.0f}" if not np.isnan(lamin_bg) else "n/a"
    lbr_bg_str = f"{lbr_bg:.0f}" if has_lbr else "n/a (no LBR)"
    dapi_bg_str = f"{dapi_bg:.0f}" if not np.isnan(dapi_bg) else "n/a"
    print(f"  done (BG: Lamin={bg_str}, LBR={lbr_bg_str}, DAPI={dapi_bg_str})")

    return [result]


def write_excel(all_results, params, save_dir):
    wb = Workbook()

    # ---- Parameters sheet ----
    ws_params = wb.active
    ws_params.title = "Parameters"
    header_font = Font(bold=True, size=11)
    ws_params["A1"] = "Parameter"
    ws_params["B1"] = "Value"
    ws_params["A1"].font = header_font
    ws_params["B1"].font = header_font

    param_rows = [
        ("Channel Lamins", params["ch_lamin"]),
        ("Channel LBR (0 = dataset has no LBR channel)", params["ch_lbr"]),
        ("Channel DAPI", params["ch_nucl"]),
        ("DAPI Threshold Method", params["threshold"]),
        ("Pixel Size (um/px)", params["pixelsize"]),
        ("LBR Shape Percentile Cutoff", params["lbr_percentile"]),
        ("LBR Edge Smoothing (sigma)", params["smooth_sigma"]),
        ("Projection", "Average intensity only"),
        ("Protein/DNA Quantification Source", "DAPI mask (Triangle threshold, projected, background-subtracted)"),
        ("Shape Source", "LBR channel, best-focus Z-plane (Laplacian variance); blank if no LBR channel"),
        ("Background Subtraction", "Mean of pixels outside DAPI mask"),
    ]
    for i, (k, v) in enumerate(param_rows, start=2):
        ws_params[f"A{i}"] = k
        ws_params[f"B{i}"] = v
    ws_params.column_dimensions["A"].width = 45
    ws_params.column_dimensions["B"].width = 55

    # ---- All Data sheet ----
    ws = wb.create_sheet("All Data", 0)

    columns = [
        "Filename", "Has LBR", "Best Z (LBR)",
        "NucPixCount", "DAPI Mask Area (um^2)",
        "BG Lamin", "BG LBR", "BG DAPI",
        "Total Lamin", "Mean Lamin",
        "Total LBR", "Mean LBR",
        "Total DAPI", "Mean DAPI",
        "Mean Lamin / Mean LBR",
        "LBR Shape Area (px)", "LBR Shape Area (um^2)",
        "LBR Shape Perimeter (um)",
        "LBR Shape Circularity", "LBR Shape Roundness",
        "LBR Shape Solidity", "LBR Shape Aspect Ratio",
        "LBR Shape Compactness",
    ]

    header_fill_blue = PatternFill("solid", fgColor="4472C4")
    header_fill_green = PatternFill("solid", fgColor="548235")
    header_fill_gray = PatternFill("solid", fgColor="808080")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(bottom=Side(style="thin"))

    lbr_shape_cols = {c for c in columns if c.startswith("LBR Shape")}
    bg_cols = {"BG Lamin", "BG LBR", "BG DAPI"}

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font_white
        if col_name in lbr_shape_cols:
            cell.fill = header_fill_green
        elif col_name in bg_cols:
            cell.fill = header_fill_gray
        else:
            cell.fill = header_fill_blue
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    for row_idx, row_data in enumerate(all_results, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))

    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for row_idx in range(2, len(all_results) + 2):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 28)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(all_results) + 1}"

    out_path = os.path.join(save_dir, "Analysis_Results.xlsx")
    wb.save(out_path)
    return out_path


# ===================== ENTRY POINT ==========================================

def main():
    data_dir, save_dir = select_folders()

    if HAS_TK:
        dlg = ParameterDialog()
        params = dlg.result
        if params is None:
            sys.exit("Cancelled.")
    else:
        params = get_params_cli()

    tiff_files = sorted(
        [f for f in glob.glob(os.path.join(data_dir, "*"))
         if f.lower().endswith((".tif", ".tiff"))]
    )

    if not tiff_files:
        print(f"No .tif/.tiff files found in {data_dir}")
        sys.exit(1)

    print(f"\nFound {len(tiff_files)} TIFF file(s). Starting analysis...\n")

    all_results = []
    for i, fpath in enumerate(tiff_files):
        fname = os.path.basename(fpath)
        print(f"[{i + 1}/{len(tiff_files)}] Processing: {fname}")
        try:
            results = process_single_image(fpath, params, save_dir)
            all_results.extend(results)
        except Exception as e:
            print(f"  ERROR processing {fname}: {e}")
            continue

    if all_results:
        xlsx_path = write_excel(all_results, params, save_dir)
        print(f"\nDone! {len(all_results)} rows written to:\n  {xlsx_path}")
    else:
        print("\nNo results were generated.")


if __name__ == "__main__":
    main()
